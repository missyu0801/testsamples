import requests
import json
import jsonpath
import openpyxl

wb = openpyxl.load_workbook("../xls_data/student.xlsx")
class Common:
    def __init__(self,filenamepath, sheetname):
        global wb
        global sheet
        wb = openpyxl.load_workbook(filenamepath)
        sheet = wb[sheetname]

    def fetch_row_count(self):
        rows = sheet.max_row
        return rows

    def fetch_colm_count(self):
        colmn = sheet.max_column
        return colmn

    def fetch_key_names(self):
        c = sheet.max_column
        li=[]
        for i in range(1,c+1):
            cell = sheet.cell(row=1, column=i)
            li.insert(i-1, cell.value)
        return li

    def update_request_with_data(self,rownumber,jsonrequest,key_list):
        c = sheet.max_column
        for i in range(1,c+1):
            cell = sheet.cell(row=rownumber, column=i)
            jsonrequest[key_list[i-1]]=cell.value
        return  jsonrequest

